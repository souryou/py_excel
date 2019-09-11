#!/usr/bin/python3
import openpyxl as xl
import tkinter as tk
from tkinter import filedialog
from tkinter.ttk import *
from tkinter import messagebox

title = "窗口标题"
window=tk.Tk()
window.title(title)
window.geometry("600x300")

fn = tk.StringVar()
startRow = tk.StringVar()
sheet_name = tk.StringVar()
tabShow = tk.StringVar()
tabSearch = tk.StringVar()

fpath = ""
sr = -1
sn = ""
nline = []
nsearch = []

e = tk.Entry(textvariable=fn)
e.place(x=0,y=0, width=550, height=35)
e.delete(0, tk.END)
e.insert(0, "选择要处理的文件")

lshow = tk.Label(window, text="显\n示\n表\n头")
lsearch = tk.Label(window, text="搜\n索\n表\n头")

lshow.place(x=0, y=70, width=20, height=100)
lsearch.place(x=0, y=170, width=20, height=100)

eshow = tk.Text()
esearch = tk.Text()
eshow.place(x=20, y=70, width=580, height=100)
esearch.place(x=20, y=170, width=580, height=100)

ldata = tk.Label(window, text="开始行数:").place(x=0, y=36, width=50, height=25)
eIndex = tk.Entry(textvariable=startRow)
eIndex.place(x=52, y=36, width=30, height=25)
eIndex.delete(0, tk.END)
eIndex.insert(0, "1")

lSName = tk.Label(window, text="sheet名称:").place(x=90, y=36, width=80, height=25)
eSName = tk.Entry(textvariable=sheet_name)
eSName.place(x=165, y=36, width=110, height=25)
eSName.delete(0, tk.END)
eSName.insert(0, "")

lprogress = tk.Label(window, text="进度:").place(x=0, y=275, width=28, height=25)

ppg = Progressbar(window, length=560, mode="determinate", orient="horizontal")
ppg.place(x=36, y=275)
ppg["value"] = 0
ppg["variable"]="hello"



def pg(cur):
	ppg["value"] = cur
	window.update()

def isOk():
	global fpath
	global sr 
	global sn
	global nline 
	global nsearch

	fpath = fn.get()
	sr = startRow.get()
	sn = sheet_name.get()
	nline = eshow.get("0.0", "end").split('@@')
	nsearch = esearch.get("0.0", "end").split('@@')
	

	if len(nline) != len(nsearch):
		messagebox.showwarning("错误", "显示表头和搜索表头个数不匹配")
		return False


	nsearch[len(nsearch)-1] = nsearch[len(nsearch)-1].replace("\r","").replace("\n","");
	try: 
		sr = int(sr)
		if sr < 1:
			messagebox.showwarning("错误", "开始行数必须是大于等于1的整数")
			return False
	except:
		messagebox.showwarning("错误", "开始行数必须是大于等于1的整数")
		return False

	if fpath != "" and fpath.endswith("xlsx"):
		buttonGo["state"]="active"
	else:
		buttonGo["state"]="disabled"

	if fpath == "" or fpath == "选择要处理的文件":
		messagebox.showwarning("错误", "文件路径不合法")
		return False
	if sn == "" :
		messagebox.showwarning("错误", "sheet名称不合法")
		return False

	return True

def filePath():
	fn = filedialog.askopenfilename()
	e.delete(0, tk.END)
	e.insert(0, fn)
	if fn != "" and fn.endswith("xlsx"):
		buttonGo["state"]="active"
	else:
		buttonGo["state"]="disabled"
		messagebox.showwarning("警告", "张老板不是说只处理2017（xlsx）的excel吗？")

def go():
	ppg["value"] = 0
	window.update()
	if isOk():
		buttonGo["state"]="disabled"
		messagebox.showinfo("提示", "处理中...")
		nn = fpath[0:fpath.rindex('.')]+"_new"+fpath[fpath.rindex('.'):]
		dwb = xl.Workbook()
		wb = xl.load_workbook(fpath, data_only=True)
		wb.guess_types = True

		ws = wb.get_sheet_by_name(sn)
		
		dws = dwb.active
		dws.title=ws.title
		
		rows = ws.max_row+1
		colums = ws.max_column+1

		ppg["maximum"] = rows - 1

		ncolumn = len(nline)+1
		noFind = []
		nindex=[]
		for ni in nsearch:
			for i in range(1, colums):
				if ws.cell(sr, i).value == ni:
					nindex.append(i)
					break
				elif (i+1)==colums:
					noFind.append(ni)
					
		if len(noFind)==0:	
			for nw in range(1, ncolumn):
				dws.cell(1, nw).value= nline[nw-1]

			for r in range(2, rows):
				for c in range(1, ncolumn):
					dws.cell(row=r, column=c).value = ws.cell(row=sr+r-1, column=nindex[c-1]).value
					pg(r)

			dwb.save(nn)
			messagebox.showinfo("提示", "恭喜，处理完成！")
			buttonGo["state"]="active"
		else:
			messagebox.showwarning("错误", "没有找到如下表头：\n"+str(noFind))
			buttonGo["state"]="active"

buttonGo=tk.Button(text="去吧，皮卡丘~", command=go)
buttonGo.place(x=300, y=36, width=300, height=35)
buttonGo["state"]="disabled"
buttonFile=tk.Button(text="选择", command=filePath).place(x=550, y=0, width=50, height=35)

tk.mainloop()
